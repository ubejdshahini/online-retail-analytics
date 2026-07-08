import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(df: pd.DataFrame, kpis: dict, recs: list[dict], projections: dict = None) -> bytes:
    """
    Generates a beautifully styled multi-sheet Excel report.
    Sheets:
      1. Executive Summary (KPIs and Simulator Projections)
      2. Customer Segments (RFM Breakdown)
      3. Actionable Recommendations

    Returns:
      bytes: In-memory Excel file bytes.
    """
    wb = Workbook()
    
    # Setup Styles
    header_fill = PatternFill(start_color="1A1D2E", end_color="1A1D2E", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    section_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="1A1D2E")
    
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    title_font = Font(name="Segoe UI", size=16, bold=True, color="6C63FF")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # ── SHEET 1: Executive Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1["A1"] = "Retail Analytics - Executive Report"
    ws1["A1"].font = title_font
    ws1.row_dimensions[1].height = 30
    
    ws1["A3"] = "Historical KPI Summary"
    ws1["A3"].font = section_font
    ws1["A3"].fill = section_fill
    ws1.merge_cells("A3:B3")
    
    kpi_rows = [
        ("Total Net Revenue", kpis.get("total_revenue", 0), "£#,##0.00"),
        ("Total Orders", kpis.get("total_orders", 0), "#,##0"),
        ("Average Order Value", kpis.get("avg_order_value", 0), "£#,##0.00"),
        ("Unique Customers", kpis.get("unique_customers", 0), "#,##0"),
        ("Top Country", kpis.get("top_country", "N/A"), None),
        ("Best Selling Product", kpis.get("best_product", "N/A"), None),
        ("Return Rate", kpis.get("return_rate_%", 0) / 100, "0.0%"),
    ]
    
    curr_row = 4
    for label, val, num_fmt in kpi_rows:
        ws1.cell(row=curr_row, column=1, value=label).font = bold_font
        cell = ws1.cell(row=curr_row, column=2, value=val)
        cell.font = regular_font
        if num_fmt:
            cell.number_format = num_fmt
        ws1.cell(row=curr_row, column=1).border = thin_border
        ws1.cell(row=curr_row, column=2).border = thin_border
        curr_row += 1
        
    if projections:
        curr_row += 1
        ws1.cell(row=curr_row, column=1, value="Simulation & Projections").font = section_font
        ws1.cell(row=curr_row, column=1).fill = section_fill
        ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
        curr_row += 1
        
        proj_rows = [
            ("VIP Retention Increase Rate", projections.get("vip_retention_pct", 0) / 100, "0%"),
            ("Win-Back Engagement Rate", projections.get("winback_rate_pct", 0) / 100, "0%"),
            ("Product Returns Reduction Rate", projections.get("return_reduction_pct", 0) / 100, "0%"),
            ("New Customer Acquisition Growth", projections.get("new_customer_growth_pct", 0) / 100, "0%"),
            ("---", "", ""), # Separator
            ("Base Revenue", projections.get("base_revenue", 0), "£#,##0.00"),
            ("VIP Retention Uplift", projections.get("revenue_from_vip", 0), "£#,##0.00"),
            ("Win-Back Revenue", projections.get("revenue_from_winback", 0), "£#,##0.00"),
            ("Returns Savings", projections.get("revenue_from_returns", 0), "£#,##0.00"),
            ("New Customer Revenue Uplift", projections.get("revenue_from_new", 0), "£#,##0.00"),
            ("Total Projected Uplift", projections.get("total_uplift", 0), "£#,##0.00"),
            ("Projected Total Revenue", projections.get("projected_revenue", 0), "£#,##0.00"),
        ]
        
        for label, val, num_fmt in proj_rows:
            if label == "---":
                curr_row += 1
                continue
            ws1.cell(row=curr_row, column=1, value=label).font = bold_font
            cell = ws1.cell(row=curr_row, column=2, value=val)
            cell.font = regular_font
            if num_fmt:
                cell.number_format = num_fmt
            ws1.cell(row=curr_row, column=1).border = thin_border
            ws1.cell(row=curr_row, column=2).border = thin_border
            
            # Draw highlight for final totals
            if label in ["Total Projected Uplift", "Projected Total Revenue"]:
                ws1.cell(row=curr_row, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="6C63FF")
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="6C63FF")
            curr_row += 1

    # ── SHEET 2: Customer Segments ──────────────────────────────────────────
    from src.recommendation_engine import compute_rfm, get_segment_summary
    rfm = compute_rfm(df)
    seg_summary = get_segment_summary(rfm)
    
    ws2 = wb.create_sheet(title="Customer Segments")
    ws2.views.sheetView[0].showGridLines = True
    
    # Title
    ws2["A1"] = "Customer Segmentation Analysis (RFM)"
    ws2["A1"].font = title_font
    ws2.row_dimensions[1].height = 30
    
    headers2 = ["Segment", "Customers Count", "Avg Recency (Days)", "Avg Frequency (Orders)", "Avg Revenue per Customer", "Total Segment Revenue"]
    ws2.append([]) # spacer
    ws2.append(headers2)
    
    # Header styling
    ws2.row_dimensions[3].height = 24
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for index, row in seg_summary.iterrows():
        r_vals = [
            row["Segment"],
            row["Customers"],
            row["Avg_Recency_Days"],
            row["Avg_Frequency"],
            row["Avg_Revenue"],
            row["Total_Revenue"]
        ]
        ws2.append(r_vals)
        r_idx = ws2.max_row
        ws2.cell(row=r_idx, column=1).font = bold_font
        ws2.cell(row=r_idx, column=2).number_format = "#,##0"
        ws2.cell(row=r_idx, column=3).number_format = "0.0"
        ws2.cell(row=r_idx, column=4).number_format = "0.0"
        ws2.cell(row=r_idx, column=5).number_format = "£#,##0.00"
        ws2.cell(row=r_idx, column=6).number_format = "£#,##0.00"
        
        # Border
        for c in range(1, 7):
            ws2.cell(row=r_idx, column=c).border = thin_border
            ws2.cell(row=r_idx, column=c).font = regular_font
            if c > 1:
                ws2.cell(row=r_idx, column=c).alignment = Alignment(horizontal="right")
            else:
                ws2.cell(row=r_idx, column=c).font = bold_font

    # ── SHEET 3: Recommendations ────────────────────────────────────────────
    ws3 = wb.create_sheet(title="Actionable Recommendations")
    ws3.views.sheetView[0].showGridLines = True
    
    # Title
    ws3["A1"] = "Strategic Recommendations"
    ws3["A1"].font = title_font
    ws3.row_dimensions[1].height = 30
    
    headers3 = ["Priority", "Type", "Target Segment", "Recommendation Title", "Estimated Profit Impact", "Action Plan Details"]
    ws3.append([]) # spacer
    ws3.append(headers3)
    ws3.row_dimensions[3].height = 24
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for r in recs:
        r_vals = [
            r.get("priority", ""),
            r.get("type", ""),
            r.get("segment", ""),
            r.get("title", ""),
            r.get("impact_pct", 0) / 100,
            r.get("message", "")
        ]
        ws3.append(r_vals)
        r_idx = ws3.max_row
        
        # Format columns
        ws3.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center")
        ws3.cell(row=r_idx, column=5).number_format = "0.0%"
        ws3.cell(row=r_idx, column=5).alignment = Alignment(horizontal="right")
        ws3.cell(row=r_idx, column=6).alignment = Alignment(wrap_text=True)
        
        # Style row
        for c in range(1, 7):
            cell = ws3.cell(row=r_idx, column=c)
            cell.border = thin_border
            cell.font = regular_font
            
        # Priority specific coloring
        p_cell = ws3.cell(row=r_idx, column=1)
        if r.get("priority") == "High":
            p_cell.font = Font(name="Segoe UI", size=10, bold=True, color="E71D36")
        elif r.get("priority") == "Medium":
            p_cell.font = Font(name="Segoe UI", size=10, bold=True, color="F7931E")
        else:
            p_cell.font = Font(name="Segoe UI", size=10, bold=True, color="2EC4B6")
            
    # Auto-adjust column widths for all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                # Skip merged or very long cells to avoid huge columns
                if cell.row == 1 or len(val_str) > 50:
                    continue
                max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Extra manual adjustment for detail columns
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25
    ws3.column_dimensions["F"].width = 65
    
    # Save to BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
